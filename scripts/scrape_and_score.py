from __future__ import annotations
"""
LinkedIn India Job Scraper & Scorer
Calls Apify API → scores jobs against user's resume → saves ranked .xlsx report

When AI_ANALYSIS=true, scoring is done by Gemini AI instead of keyword matching.
"""

import os
import re
import json
import time
import requests
from datetime import datetime, date, timedelta
from collections import Counter
from dotenv import load_dotenv

load_dotenv()  # Load .env file for local runs (GitHub Actions injects secrets directly)

# Initialize run logger — only when running as main script
# (scrape_naukri.py imports from this file, so we must not overwrite its logger)
from run_logger import new_logger, get_logger
_logger = None

# AI analysis toggle
from ai_analyzer import is_ai_enabled, analyze_jobs as ai_analyze_jobs

# ── openpyxl (installed via requirements.txt) ────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═════════════════════════════════════════════════════════════════════
# CONFIG  (values injected from GitHub Secrets)
# ═════════════════════════════════════════════════════════════════════
from apify_token import get_apify_token, try_next_token, is_auth_error, get_current_token_name

ACTOR_ID      = "curious_coder~linkedin-jobs-scraper"
TODAY         = date.today().isoformat()          # e.g. "2026-02-22"
YESTERDAY     = (date.today() - timedelta(days=1)).isoformat()

# ── LinkedIn search URLs ─────────────────────────────────────────────
# Configurable via SEARCH_KEYWORDS and SEARCH_LOCATION env vars.
# If not set, falls back to default URLs below.
def _build_linkedin_urls() -> list[str]:
    """Build LinkedIn search URLs from env vars or use defaults."""
    keywords_raw = os.environ.get("SEARCH_KEYWORDS", "").strip()
    location = os.environ.get("SEARCH_LOCATION", "India").strip()

    if keywords_raw:
        # User provided custom search keywords (comma-separated)
        keywords_list = [k.strip() for k in keywords_raw.split(",") if k.strip()]
        urls = []
        for kw in keywords_list:
            encoded = kw.replace(" ", "%20")
            urls.append(
                f"https://www.linkedin.com/jobs/search/?keywords={encoded}"
                f"&location={location.replace(' ', '%20')}"
                f"&f_TPR=r86400&position=1&pageNum=0"
            )
        return urls

    # Default: uses hardcoded URLs (Node.js focused)
    return [
        "https://www.linkedin.com/jobs/search/?keywords=Backend%20Developer%20Node.js&location=India&f_TPR=r86400&position=1&pageNum=0",
        "https://www.linkedin.com/jobs/search/?keywords=MERN%20Stack%20Developer&location=India&f_TPR=r86400&position=1&pageNum=0",
        "https://www.linkedin.com/jobs/search/?keywords=Full%20Stack%20Developer%20Node%20React&location=India&f_TPR=r86400&position=1&pageNum=0",
        "https://www.linkedin.com/jobs/search/?keywords=Node.js%20Microservices%20Developer&location=India&f_TPR=r86400&position=1&pageNum=0",
        "https://www.linkedin.com/jobs/search/?keywords=Software%20Engineer%20Node.js%20Backend&location=India&f_TPR=r86400&position=1&pageNum=0",
    ]

SEARCH_URLS = _build_linkedin_urls()

# ═════════════════════════════════════════════════════════════════════
# SKILL TAXONOMY  (customise for your resume)
# ═════════════════════════════════════════════════════════════════════
MUST_HAVE = [
    "node.js","nodejs","node js","express","mongodb","react","typescript",
    "javascript","mern","rest api","restful api","restful",
]
BACKEND_PLUS = [
    "microservices","redis","jwt","docker","aws","kubernetes","lambda","s3",
    "sqs","sns","api gateway","nosql","mongoose","nestjs","nest.js","graphql",
    "postgresql","mysql","kafka","rabbitmq","oauth","rbac","ldap","azure","gcp",
    "serverless","authentication","eventbridge","cdk",
]
NICE_TO_HAVE = [
    "ci/cd","jest","git","agile","scalable","cloud","performance","caching",
    "websocket","socket.io","prisma","terraform","elk","prometheus","grafana",
    "langchain","rag","llm","ai","crewai","vector","langsmith","mcp",
]

LABEL = {
    "node.js":"Node.js","nodejs":"Node.js","node js":"Node.js",
    "express":"Express.js","mongodb":"MongoDB","react":"React.js",
    "typescript":"TypeScript","javascript":"JavaScript","mern":"MERN",
    "rest api":"REST API","restful api":"REST API","restful":"REST API",
    "microservices":"Microservices","redis":"Redis","jwt":"JWT",
    "docker":"Docker","aws":"AWS","kubernetes":"Kubernetes",
    "lambda":"AWS Lambda","authentication":"Authentication",
    "nestjs":"NestJS","nest.js":"NestJS","graphql":"GraphQL",
    "postgresql":"PostgreSQL","mysql":"MySQL","kafka":"Kafka",
    "rabbitmq":"RabbitMQ","nosql":"NoSQL","mongoose":"Mongoose",
    "azure":"Azure","gcp":"GCP","serverless":"Serverless",
    "oauth":"OAuth","rbac":"RBAC","ldap":"LDAP","ci/cd":"CI/CD",
    "jest":"Jest","caching":"Caching","websocket":"WebSocket",
    "prisma":"Prisma","terraform":"Terraform","s3":"AWS S3",
    "sqs":"AWS SQS","sns":"AWS SNS","api gateway":"API Gateway",
    "eventbridge":"EventBridge","cdk":"AWS CDK",
    "langchain":"LangChain","rag":"RAG","llm":"LLM",
    "ai":"AI/ML","crewai":"CrewAI","vector":"Vector DB","langsmith":"LangSmith","mcp":"MCP",
}

# ═════════════════════════════════════════════════════════════════════
# APIFY  —  run actor & fetch results
# ═════════════════════════════════════════════════════════════════════
def run_apify_actor() -> list[dict]:
    """Trigger the Apify actor and poll until done, then return dataset items.
    Cascades through APIFY_TOKEN_1, _2, _3... on auth failure."""
    token = get_apify_token()
    while True:
        result = _try_apify_run(token)
        if result is not None:
            return result
        # Auth failed — try next token
        token = try_next_token("scrape_and_score.py (LinkedIn)")
        if not token:
            raise RuntimeError("All Apify tokens exhausted. Cannot scrape LinkedIn.")


def _try_apify_run(token: str) -> list[dict] | None:
    """Attempt an Apify actor run. Returns None on auth failure, list on success."""
    base = "https://api.apify.com/v2"
    params = {"token": token}

    print(f"\n▶  Starting LinkedIn actor … ({get_current_token_name()})")
    run_resp = requests.post(
        f"{base}/acts/{ACTOR_ID}/runs",
        params=params,
        json={
            "urls":  SEARCH_URLS,
            "count": 100,
            "scrapeCompany": False,
        },
        timeout=30,
    )

    if is_auth_error(run_resp.status_code):
        print(f"   ⚠️  {get_current_token_name()} failed (HTTP {run_resp.status_code})")
        return None

    run_resp.raise_for_status()
    run_id = run_resp.json()["data"]["id"]
    print(f"   Run ID: {run_id}")

    # Poll until finished (max 10 min)
    for attempt in range(60):
        time.sleep(10)
        status_resp = requests.get(
            f"{base}/actor-runs/{run_id}", params=params, timeout=15
        )
        status = status_resp.json()["data"]["status"]
        print(f"   [{attempt+1:02d}] Status: {status}")
        if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            break

    if status != "SUCCEEDED":
        raise RuntimeError(f"Actor run did not succeed — final status: {status}")

    dataset_id = status_resp.json()["data"]["defaultDatasetId"]
    items_resp  = requests.get(
        f"{base}/datasets/{dataset_id}/items",
        params={**params, "format": "json", "limit": 500},
        timeout=30,
    )
    items_resp.raise_for_status()
    raw = items_resp.json()
    print(f"   Fetched {len(raw)} raw items from dataset.")
    return raw


# ═════════════════════════════════════════════════════════════════════
# SCORING
# ═════════════════════════════════════════════════════════════════════
def extract_exp(desc: str) -> tuple[str, int, int]:
    m = re.search(r"(\d+)\s*[-–to]+\s*(\d+)\s*years?", desc)
    if m:
        return f"{m.group(1)}–{m.group(2)} yrs", int(m.group(1)), int(m.group(2))
    m2 = re.search(r"(\d+)\+\s*years?", desc)
    if m2:
        n = int(m2.group(1))
        return f"{n}+ yrs", n, n + 3
    return "Not specified", 0, 10


def score_job(job: dict) -> dict:
    desc = (job.get("descriptionText","") + " " + job.get("title","")).lower()
    loc  = job.get("location","").lower()

    p1 = sum(1 for s in MUST_HAVE    if s in desc)
    p2 = sum(1 for s in BACKEND_PLUS if s in desc)
    p3 = sum(1 for s in NICE_TO_HAVE if s in desc)

    exp_str, lo, hi = extract_exp(desc)
    exp_ok     = (lo <= 4 and hi >= 2)
    is_intern  = any(w in desc for w in ["intern","internship","fresher","graduate trainee"])
    is_senior  = any(w in desc for w in ["10+ years","9+ years","8+ years","staff engineer","principal engineer"])
    is_remote  = "remote" in loc or "remote" in desc[:600]

    score = (
        p1 * 15 + p2 * 5 + p3 * 2
        + (12 if exp_ok    else 0)
        + (6  if is_remote else 0)
        - (25 if is_intern  else 0)
        - (8  if is_senior  else 0)
    )
    max_s = len(MUST_HAVE)*15 + len(BACKEND_PLUS)*5 + len(NICE_TO_HAVE)*2 + 12 + 6
    pct   = max(10, min(round(score / max_s * 100), 97))
    prob  = "High" if pct >= 65 else ("Medium" if pct >= 48 else "Low")

    # Key skills (deduped pretty labels)
    found, seen_l = [], set()
    for s in MUST_HAVE + BACKEND_PLUS + NICE_TO_HAVE:
        if s in desc:
            lbl = LABEL.get(s, s.title())
            if lbl not in seen_l:
                seen_l.add(lbl)
                found.append(lbl)

    # Work mode
    if   "fully remote" in desc[:800]: wmode = "Fully Remote"
    elif "remote" in loc:              wmode = "Fully Remote"
    elif "hybrid" in desc[:400]:       wmode = "Hybrid"
    elif "remote" in desc[:400]:       wmode = "Remote"
    else:                              wmode = "Onsite"

    # Why match
    reasons = []
    if any(x in desc for x in ["node.js","nodejs","node js"]): reasons.append("Core Node.js backend role")
    if "react"    in desc:              reasons.append("React.js front-end valued")
    if "mongodb"  in desc:              reasons.append("MongoDB/Mongoose match")
    if "aws"      in desc:              reasons.append("AWS/Lambda cloud experience relevant")
    if "microservices" in desc:         reasons.append("Microservices architecture required")
    if "redis"    in desc:              reasons.append("Redis caching expertise needed")
    if "jwt"      in desc or "oauth" in desc: reasons.append("JWT/OAuth security experience valued")
    if "docker"   in desc:              reasons.append("Docker containerisation required")
    if "nestjs"   in desc or "nest.js" in desc: reasons.append("NestJS framework match")
    if any(x in desc for x in ["langchain","rag","llm","crewai"]): reasons.append("AI/LLM skills are a differentiator")
    why = "; ".join(reasons[:3]) if reasons else "Full-stack Node.js/REST API alignment"

    return {
        "title":        job.get("title","").strip(),
        "company":      job.get("companyName","").strip(),
        "location":     job.get("location","").strip() or "India",
        "work_mode":    wmode,
        "posted":       job.get("postedAt",""),
        "exp_req":      exp_str,
        "key_skills":   ", ".join(found[:10]),
        "why_match":    why,
        "link":         job.get("link",""),
        "match_score":  pct,
        "interview_prob": prob,
        "score_raw":    score,
    }


def process_jobs(raw: list[dict]) -> list[dict]:
    jobs = []
    for item in raw:
        # Actor returns items directly as dicts with job fields
        if isinstance(item, dict) and "title" in item:
            jobs.append(item)

    print(f"   Valid job records: {len(jobs)}")

    # Filter last 24 h
    recent = [j for j in jobs if j.get("postedAt","") in (TODAY, YESTERDAY)]
    print(f"   Jobs in last 24 h: {len(recent)}")
    if not recent:
        recent = jobs   # fallback — use all if date filter too strict

    scored = [score_job(j) for j in recent]
    scored.sort(key=lambda x: x["score_raw"], reverse=True)

    # Deduplicate
    seen, unique = set(), []
    for j in scored:
        key = (j["title"].lower()[:40], j["company"].lower())
        if key not in seen:
            seen.add(key)
            unique.append(j)

    return unique[:20]


# ═════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ═════════════════════════════════════════════════════════════════════
NAVY      = "0D2137"
TEAL      = "0E7C7B"
HEADER_R  = "1A3A5C"
HIGH_BG   = "D4EDDA"; HIGH_FG = "155724"
MED_BG    = "FFF3CD"; MED_FG  = "856404"
LOW_BG    = "F8D7DA"; LOW_FG  = "721C24"
ROW_ALT   = "EEF4FB"; ROW_WHT = "FFFFFF"
BORDER_C  = "C9D6E3"

def _bdr():
    t = Side(style="thin", color=BORDER_C)
    return Border(left=t, right=t, top=t, bottom=t)


def build_excel(jobs: list[dict], out_path: str, platform: str = "LinkedIn", ai_mode: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "🇮🇳 India Job Matches"

    # ── Column config (changes based on AI mode) ────────────────────
    if ai_mode:
        HEADERS = [
            "#","Status","Job Title","Company Name","Location","Work Mode",
            "Posted Date","Exp Required","Key Skills Required",
            "Match\nScore (%)","Role Fit","AI Summary",
            "Missing Skills","Resume Tips","Job Link",
        ]
        COL_W = [4,13,30,22,20,13,13,13,38,12,14,44,36,36,18]
        LAST_COL = "O"
        CENTER_COLS = {1,2,6,7,8,10,11}
        SCORE_COL = 10
        FIT_COL   = 11
        LINK_COL  = 15
        STATUS_COL = 2
    else:
        HEADERS = [
            "#","Job Title","Company Name","Location","Work Mode",
            "Posted Date","Exp Required","Key Skills Required",
            "Why It Matches My Resume","Match\nScore (%)","Interview\nProbability","Job Link",
        ]
        COL_W = [4,32,24,22,14,13,14,46,46,12,14,18]
        LAST_COL = "L"
        CENTER_COLS = {1,5,6,7,10,11}
        SCORE_COL = 10
        FIT_COL   = 11
        LINK_COL  = 12

    num_cols = len(HEADERS)

    # ── Banner ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    ai_tag = "  🤖 AI-Powered" if ai_mode else ""
    c.value = f"🇮🇳  {platform} India — Daily Job Match Report{ai_tag}  |  {TODAY}"
    c.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # User info banner (dynamic from config)
    from config import get_config
    cfg = get_config()
    user_label = cfg.user_name if cfg.user_name else "Your Profile"
    ws.merge_cells(f"A2:{LAST_COL}2")
    s = ws["A2"]
    s.value = f"{user_label}  |  AI-Powered Job Matching  |  See 'Scoring Guide' sheet for methodology"
    s.font  = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    s.fill  = PatternFill("solid", fgColor=TEAL)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Headers ──────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 34
    for col,(h,w) in enumerate(zip(HEADERS,COL_W),1):
        c = ws.cell(row=3,column=col,value=h)
        c.font      = Font(name="Arial",bold=True,size=9,color="FFFFFF")
        c.fill      = PatternFill("solid",fgColor=HEADER_R)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border    = _bdr()
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Data ─────────────────────────────────────────────────────────
    for i,job in enumerate(jobs,4):
        bg = ROW_ALT if i % 2 == 0 else ROW_WHT
        ws.row_dimensions[i].height = 70 if ai_mode else 60

        if ai_mode:
            vals = [
                i-3, job.get("job_status", "✨ NEW"),
                job["title"], job["company"], job["location"],
                job["work_mode"], job["posted"], job["exp_req"],
                job["key_skills"],
                job["match_score"], job["interview_prob"],
                job.get("ai_summary", "—"),
                job.get("missing_skills", "—"),
                job.get("resume_tips", "—"),
                "🔗 View Job",
            ]
        else:
            vals = [
                i-3, job["title"], job["company"], job["location"],
                job["work_mode"], job["posted"], job["exp_req"],
                job["key_skills"], job["why_match"],
                job["match_score"], job["interview_prob"], "🔗 View Job",
            ]

        for col,val in enumerate(vals,1):
            c = ws.cell(row=i,column=col,value=val)
            c.border    = _bdr()
            c.alignment = Alignment(
                vertical="top", wrap_text=True,
                horizontal="center" if col in CENTER_COLS else "left",
            )

            if col == SCORE_COL:    # Match Score
                score = job["match_score"]
                c.value = f"{score}%"
                if score >= 65:
                    c.fill = PatternFill("solid",fgColor=HIGH_BG)
                    c.font = Font(name="Arial",size=10,bold=True,color=HIGH_FG)
                elif score >= 48:
                    c.fill = PatternFill("solid",fgColor=MED_BG)
                    c.font = Font(name="Arial",size=10,bold=True,color=MED_FG)
                else:
                    c.fill = PatternFill("solid",fgColor="E3EEF9")
                    c.font = Font(name="Arial",size=10,bold=True,color="1A5276")

            elif col == FIT_COL:    # Interview Probability / Role Fit
                prob = job["interview_prob"]
                if prob in ("High", "Strong Fit"):
                    c.fill = PatternFill("solid",fgColor=HIGH_BG)
                    c.font = Font(name="Arial",size=9,bold=True,color=HIGH_FG)
                elif prob in ("Medium", "Moderate Fit"):
                    c.fill = PatternFill("solid",fgColor=MED_BG)
                    c.font = Font(name="Arial",size=9,bold=True,color=MED_FG)
                else:
                    c.fill = PatternFill("solid",fgColor=LOW_BG)
                    c.font = Font(name="Arial",size=9,bold=True,color=LOW_FG)

            elif col == LINK_COL:   # Hyperlink
                if job.get("link"):
                    c.hyperlink = job["link"]
                    c.font = Font(name="Arial",size=9,color="0563C1",underline="single")
                c.fill = PatternFill("solid",fgColor=bg)

            elif col == 1:
                c.font = Font(name="Arial",size=9,bold=True,color=NAVY)
                c.fill = PatternFill("solid",fgColor=bg)

            elif ai_mode and col == STATUS_COL:
                status_val = str(val)
                if "NEW" in status_val:
                    c.fill = PatternFill("solid", fgColor="E8F5E9")
                    c.font = Font(name="Arial", size=9, bold=True, color="2E7D32")
                else:
                    c.fill = PatternFill("solid", fgColor="F5F5F5")
                    c.font = Font(name="Arial", size=8, color="757575")
            else:
                c.font = Font(name="Arial",size=9)
                c.fill = PatternFill("solid",fgColor=bg)

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{LAST_COL}{3+len(jobs)}"

    # ── Footer ───────────────────────────────────────────────────────
    fr = 4 + len(jobs)
    ws.merge_cells(f"A{fr}:{LAST_COL}{fr}")
    fc = ws[f"A{fr}"]
    scoring_method = "Gemini AI analysis" if ai_mode else "keyword alignment with resume skills"
    fc.value = (f"Auto-generated by {platform} India Job Tracker · Run date: {TODAY} 07:00 IST  "
                f"· Sorted by Match Score (highest first) · Scores based on {scoring_method}")
    fc.font  = Font(name="Arial",size=8,italic=True,color="666666")
    fc.fill  = PatternFill("solid",fgColor="F0F4F8")
    fc.alignment = Alignment(wrap_text=True,horizontal="left",vertical="center")
    ws.row_dimensions[fr].height = 22

    # ── Sheet 2: Scoring Guide ────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Scoring Guide")
    ws2.column_dimensions["A"].width = 80

    if ai_mode:
        guide = [
            ("AI-POWERED MATCH SCORE METHODOLOGY", True, 12, "FFFFFF", NAVY),
            ("Scores are generated by Gemini AI, which reads your full resume profile and compares it against each job posting.", False, 9, "222222", "FFFFFF"),
            ("The AI considers: skill overlap, experience level, tech stack alignment, seniority fit, and role responsibilities.", False, 9, "222222", "FFFFFF"),
            ("", False, 9, "000000", "FFFFFF"),
            ("ROLE FIT CATEGORIES", True, 12, "FFFFFF", NAVY),
            ("🟢  Strong Fit (80-100%) — Excellent alignment; apply immediately", False, 9, HIGH_FG, HIGH_BG),
            ("🟡  Moderate Fit (60-79%) — Good match with some gaps; tailor resume", False, 9, MED_FG, MED_BG),
            ("🟠  Weak Fit (40-59%) — Partial match; address missing skills in cover letter", False, 9, "8B0000", LOW_BG),
            ("🔴  Not a Fit (0-39%) — Different stack/seniority; skip unless very interested", False, 9, "8B0000", LOW_BG),
            ("", False, 9, "000000", "FFFFFF"),
            ("AI COLUMNS EXPLAINED", True, 12, "FFFFFF", NAVY),
            ("AI Summary: 2-3 sentence explanation of why this job is/isn't a good match", False, 9, "222222", "FFFFFF"),
            ("Missing Skills: Skills the job requires that you don't have (yet)", False, 9, "222222", "FFFFFF"),
            ("Resume Tips: Specific, actionable suggestions to tailor your resume for this role", False, 9, "222222", "FFFFFF"),
        ]
    else:
        guide = [
            ("MATCH SCORE METHODOLOGY", True, 12, "FFFFFF", NAVY),
            ("Primary Skills (15 pts each): Node.js, Express.js, MongoDB, React.js, TypeScript, JavaScript, MERN, REST API", False, 9, "222222", "FFFFFF"),
            ("Backend+ Skills (5 pts each): Microservices, Redis, JWT, Docker, AWS, Kubernetes, Lambda, NestJS, GraphQL, PostgreSQL, Kafka, OAuth, RBAC, LDAP, Azure, GCP, Serverless, EventBridge, CDK", False, 9, "222222", "FFFFFF"),
            ("Nice-to-Have (2 pts each): CI/CD, Jest, Git, Agile, Scalable, Caching, WebSocket, Prisma, Terraform, LangChain, RAG, LLM, CrewAI, Vector DB, LangSmith, MCP", False, 9, "222222", "FFFFFF"),
            ("Experience 2–5 yrs: +12 pts  |  Remote confirmed: +6 pts", False, 9, HIGH_FG, HIGH_BG),
            ("Intern/Fresher role: −25 pts  |  Over-senior (8+ yrs): −8 pts", False, 9, "8B0000", LOW_BG),
            ("", False, 9, "000000", "FFFFFF"),
            ("INTERVIEW PROBABILITY", True, 12, "FFFFFF", NAVY),
            ("🟢  High   (≥ 65%) — Strong overlap; apply immediately", False, 9, HIGH_FG, HIGH_BG),
            ("🟡  Medium (48–64%) — Good match; tailor resume keywords", False, 9, MED_FG, MED_BG),
            ("🔴  Low    (< 48%) — Partial match; customise cover letter", False, 9, "8B0000", LOW_BG),
        ]

    for r,(text,bold,size,color,bg) in enumerate(guide,1):
        c = ws2.cell(row=r,column=1,value=text)
        c.font = Font(name="Arial",bold=bold,size=size,color=color)
        c.fill = PatternFill("solid",fgColor=bg)
        c.alignment = Alignment(wrap_text=True,vertical="center")
        ws2.row_dimensions[r].height = 8 if not text else 18

    wb.save(out_path)
    print(f"✅  Excel report saved → {out_path}")


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    _logger = new_logger(section="LinkedIn Scraper")
    ai_mode = is_ai_enabled()
    raw  = run_apify_actor()

    # Log Apify credit usage after scraping
    from apify_token import log_apify_usage
    log_apify_usage()

    if ai_mode:
        print("\n🤖  AI_ANALYSIS is ON — using Gemini AI for scoring.")
        # Process raw items into valid job records (same filtering)
        valid_jobs = []
        for item in raw:
            if isinstance(item, dict) and "title" in item:
                valid_jobs.append(item)
        print(f"   Valid job records: {len(valid_jobs)}")

        # Filter last 24 h
        recent = [j for j in valid_jobs if j.get("postedAt","") in (TODAY, YESTERDAY)]
        print(f"   Jobs in last 24 h: {len(recent)}")
        if not recent:
            recent = valid_jobs

        # Deduplicate before sending to AI (save API calls)
        seen, unique_raw = set(), []
        for j in recent:
            key = (j.get("title","").lower()[:40], j.get("companyName","").lower())
            if key not in seen:
                seen.add(key)
                unique_raw.append(j)
        job_count = int(os.environ.get("JOB_COUNT", "20"))
        unique_raw = unique_raw[:job_count]  # cap to save API calls

        jobs = ai_analyze_jobs(unique_raw, platform="LinkedIn")
    else:
        print("\n📊  AI_ANALYSIS is OFF — using keyword-based scoring.")
        jobs = process_jobs(raw)

    if not jobs:
        print("⚠️  No matching jobs found today. Exiting.")
        exit(0)

    print(f"\n📋  Top {len(jobs)} jobs selected.")
    for i,j in enumerate(jobs,1):
        status = j.get('job_status', '✨ NEW')[:12]
        print(f"  {i:2}. {j['match_score']:2}% | {j['interview_prob']:>15} | {status:<12} | {j['title'][:40]} | {j['company'][:25]}")

    out_path = f"reports/LinkedIn_India_Jobs_{TODAY}.xlsx"
    os.makedirs("reports", exist_ok=True)
    build_excel(jobs, out_path, ai_mode=ai_mode)

    # Write summary for email step
    mode_tag = "🤖 AI" if ai_mode else "📊 Keyword"
    summary_lines = [
        f"📅 Date: {TODAY}",
        f"📊 Jobs scraped & ranked: {len(jobs)} ({mode_tag} scoring)",
        "",
        "Top 5 Matches:",
    ]
    for i,j in enumerate(jobs[:5],1):
        summary_lines.append(f"  {i}. [{j['match_score']}% | {j['interview_prob']}] {j['title']} @ {j['company']} ({j['work_mode']})")
    summary_lines += ["", "Full report attached as .xlsx ✅"]

    with open("reports/email_summary.txt","w") as f:
        f.write("\n".join(summary_lines))

    print("\n📧  Email summary written.")
    print("\n".join(summary_lines))

    # Always save run log
    _logger.save()
